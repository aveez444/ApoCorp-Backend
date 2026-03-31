# apps/reports/engine.py
#
# The report engine.  Takes a validated config dict and returns either:
#   - a list of row dicts  (for JSON response)
#   - an openpyxl Workbook (for Excel download)
#
# All queries are anchored on the Enquiry model so every row represents
# one enquiry, with joined data from the other modules.

from __future__ import annotations

import datetime
from decimal import Decimal

from django.db.models import Q

from apps.enquiries.models import Enquiry
from .field_registry import FIELD_REGISTRY, MODULE_JOINS, get_field_def


# ─────────────────────────────────────────────────────────────────────────────
# Query builder
# ─────────────────────────────────────────────────────────────────────────────

class ReportEngine:

    def __init__(self, config: dict, tenant):
        self.config  = config
        self.tenant  = tenant
        self.modules  = config.get("modules", ["enquiry"])
        self.columns  = config.get("columns", [])   # [{ module, field, label }]
        self.filters  = config.get("filters",  [])  # [{ module, field, operator, value }]
        self.order_by = config.get("order_by", "-created_at")

        # Always include enquiry — it's the root
        if "enquiry" not in self.modules:
            self.modules = ["enquiry"] + self.modules

    # ── Public ────────────────────────────────────────────────────────────────

    def get_rows(self, page: int = 1, page_size: int = 50) -> dict:
        """Return paginated rows + meta."""
        qs     = self._build_queryset()
        total  = qs.count()
        offset = (page - 1) * page_size
        rows   = list(qs[offset: offset + page_size])

        columns = self._resolved_columns()
        data    = [self._serialize_row(r, columns) for r in rows]

        return {
            "columns": [{"key": c["key"], "label": c["label"]} for c in columns],
            "rows":    data,
            "total":   total,
            "page":    page,
            "pages":   max((total + page_size - 1) // page_size, 1),
            "has_next": page * page_size < total,
        }

    def get_workbook(self):
        """Return an openpyxl Workbook with all rows (no pagination)."""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            raise ImportError("openpyxl is required for Excel export. Run: pip install openpyxl")

        qs      = self._build_queryset()
        columns = self._resolved_columns()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Report"

        # ── Header row styling ──────────────────────────────────────────────
        header_font    = Font(bold=True, color="FFFFFF")
        header_fill    = PatternFill(fill_type="solid", fgColor="1E3A5F")
        header_align   = Alignment(horizontal="center", vertical="center")

        for col_idx, col in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col["label"])
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = header_align
            ws.column_dimensions[
                openpyxl.utils.get_column_letter(col_idx)
            ].width = max(len(col["label"]) + 4, 14)

        # ── Data rows ───────────────────────────────────────────────────────
        alt_fill = PatternFill(fill_type="solid", fgColor="F0F4F8")

        for row_idx, raw_row in enumerate(qs, start=2):
            row_data = self._serialize_row(raw_row, columns)
            fill     = alt_fill if row_idx % 2 == 0 else None

            for col_idx, col in enumerate(columns, start=1):
                value = row_data.get(col["key"])
                cell  = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = Alignment(vertical="center")
                if fill:
                    cell.fill = fill

        # Freeze header
        ws.freeze_panes = "A2"

        return wb

    # ── Private ───────────────────────────────────────────────────────────────

    def _build_queryset(self):
        # Start with tenant-scoped Enquiry queryset
        qs = Enquiry.objects.filter(tenant=self.tenant, is_active=True)

        # Add select_related for every included module
        joins = set()
        for mod in self.modules:
            for j in MODULE_JOINS.get(mod, []):
                joins.add(j)
        if joins:
            qs = qs.select_related(*joins)

        # Apply filters
        q_obj = self._build_filter_q()
        if q_obj:
            qs = qs.filter(q_obj)

        # Apply ordering
        try:
            qs = qs.order_by(self.order_by)
        except Exception:
            qs = qs.order_by("-created_at")

        return qs

    def _build_filter_q(self) -> Q | None:
        combined = Q()
        for f in self.filters:
            mod      = f.get("module")
            field    = f.get("field")
            operator = f.get("operator", "eq")
            value    = f.get("value")

            defn = get_field_def(mod, field)
            if not defn:
                continue  # skip unknown fields silently

            orm_path = defn["path"]
            value    = self._coerce_value(value, defn["type"], operator)

            try:
                if operator == "eq":
                    combined &= Q(**{orm_path: value})
                elif operator == "neq":
                    combined &= ~Q(**{orm_path: value})
                elif operator == "in":
                    if isinstance(value, list) and value:
                        combined &= Q(**{f"{orm_path}__in": value})
                elif operator == "gte":
                    combined &= Q(**{f"{orm_path}__gte": value})
                elif operator == "lte":
                    combined &= Q(**{f"{orm_path}__lte": value})
                elif operator == "contains":
                    combined &= Q(**{f"{orm_path}__icontains": value})
                elif operator == "isnull":
                    combined &= Q(**{f"{orm_path}__isnull": bool(value)})
            except Exception:
                continue  # bad filter value — skip gracefully

        return combined if combined.children else None

    def _resolved_columns(self) -> list[dict]:
        """
        Return a flat list of column defs in the order the user specified.
        Each entry: { key, label, path, type }
        Key is unique — module__field.
        """
        resolved = []
        seen     = set()

        for col in self.columns:
            mod   = col.get("module")
            field = col.get("field")
            label = col.get("label", "")

            defn = get_field_def(mod, field)
            if not defn:
                continue

            key = f"{mod}__{field}"
            if key in seen:
                continue
            seen.add(key)

            resolved.append({
                "key":   key,
                "label": label or defn["label"],
                "path":  defn["path"],
                "type":  defn["type"],
            })

        # Fallback: if no columns selected, show enquiry defaults
        if not resolved:
            for field_key, defn in FIELD_REGISTRY["enquiry"].items():
                resolved.append({
                    "key":   f"enquiry__{field_key}",
                    "label": defn["label"],
                    "path":  defn["path"],
                    "type":  defn["type"],
                })

        return resolved

    def _serialize_row(self, obj, columns: list[dict]) -> dict:
        """
        Walk a model instance and extract each column value via attribute
        traversal (matches ORM __ path → dot path on Python objects).
        """
        row = {}
        for col in columns:
            path  = col["path"]
            value = self._traverse(obj, path)
            value = self._format_value(value, col["type"])
            row[col["key"]] = value
        return row

    @staticmethod
    def _traverse(obj, orm_path: str):
        """
        Follow a __ separated ORM path on a model instance.
        e.g. "customer__company_name" → obj.customer.company_name
        """
        parts = orm_path.split("__")
        value = obj
        for part in parts:
            if value is None:
                return None
            value = getattr(value, part, None)
        # Handle callables (e.g. get_status_display)
        if callable(value):
            value = value()
        return value

    @staticmethod
    def _format_value(value, field_type: str):
        """Coerce value to a JSON-safe / Excel-safe Python type."""
        if value is None:
            return None
        if field_type in ("date",) and isinstance(value, (datetime.date, datetime.datetime)):
            return str(value.date() if isinstance(value, datetime.datetime) else value)
        if field_type == "datetime" and isinstance(value, datetime.datetime):
            # Return ISO string — friendly for both JSON and Excel
            return value.strftime("%Y-%m-%d %H:%M")
        if isinstance(value, Decimal):
            return float(value)
        if isinstance(value, bool):
            return value
        return value

    @staticmethod
    def _coerce_value(value, field_type: str, operator: str):
        """Parse user-supplied filter values into correct Python types."""
        if operator == "in":
            return value if isinstance(value, list) else [value]

        if field_type in ("date", "datetime") and isinstance(value, str):
            try:
                return datetime.date.fromisoformat(value)
            except ValueError:
                return value

        if field_type == "decimal" and isinstance(value, (str, int, float)):
            try:
                return Decimal(str(value))
            except Exception:
                return value

        if field_type == "int" and isinstance(value, str):
            try:
                return int(value)
            except ValueError:
                return value

        return value