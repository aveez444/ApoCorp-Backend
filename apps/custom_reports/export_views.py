# apps/reports/export_views.py

import io
import datetime
from decimal import Decimal
from typing import Any, Dict, List

from django.http import HttpResponse
from rest_framework import status
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    raise ImportError("openpyxl is required for Excel export. Run: pip install openpyxl")


class ExcelExporter:
    """Universal Excel exporter with proper formatting."""
    
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    HEADER_FILL = PatternFill(fill_type="solid", fgColor="1E3A5F")
    HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
    CELL_ALIGN = Alignment(vertical="center", wrap_text=True)
    NUMBER_ALIGN = Alignment(horizontal="right", vertical="center")
    DATE_ALIGN = Alignment(horizontal="left", vertical="center")
    CENTER_ALIGN = Alignment(horizontal="center", vertical="center")
    ALT_ROW_FILL = PatternFill(fill_type="solid", fgColor="F0F4F8")
    BORDER = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )
    
    def __init__(self, title: str = "Report"):
        self.wb = openpyxl.Workbook()
        self.ws = self.wb.active
        self.ws.title = title[:31]
        self.current_row = 1
        self.column_formats = {}
        
    def add_header(self, columns: List[Dict[str, Any]]) -> None:
        for col_idx, col in enumerate(columns, start=1):
            cell = self.ws.cell(row=self.current_row, column=col_idx, 
                                value=col.get("label", col.get("key", "")))
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = self.HEADER_ALIGN
            cell.border = self.BORDER
            
            col_letter = get_column_letter(col_idx)
            width = max(len(str(cell.value)) + 4, 12)
            self.ws.column_dimensions[col_letter].width = min(width, 50)
            self.column_formats[col_idx] = col.get("type", "str")
        
        self.current_row += 1
    
    def add_row(self, row_data: Dict[str, Any], columns: List[Dict[str, Any]], 
                is_alternate: bool = False) -> None:
        for col_idx, col in enumerate(columns, start=1):
            key = col.get("key")
            value = row_data.get(key)
            field_type = col.get("type", "str")
            
            formatted_value = self._format_cell_value(value, field_type)
            cell = self.ws.cell(row=self.current_row, column=col_idx, value=formatted_value)
            cell.border = self.BORDER
            
            if field_type in ("decimal", "int", "float", "number"):
                cell.alignment = self.NUMBER_ALIGN
                if isinstance(formatted_value, (int, float)):
                    cell.number_format = '#,##0.00' if field_type == "decimal" else '#,##0'
            elif field_type in ("date", "datetime"):
                cell.alignment = self.DATE_ALIGN
                if field_type == "date":
                    cell.number_format = 'yyyy-mm-dd'
                else:
                    cell.number_format = 'yyyy-mm-dd hh:mm'
            elif field_type == "bool":
                cell.alignment = self.CENTER_ALIGN
            else:
                cell.alignment = self.CELL_ALIGN
            
            if is_alternate:
                cell.fill = self.ALT_ROW_FILL
        
        self.current_row += 1
    
    def _format_cell_value(self, value: Any, field_type: str) -> Any:
        if value is None:
            return ""
        
        if field_type == "date":
            if isinstance(value, datetime.datetime):
                return value.date()
            elif isinstance(value, datetime.date):
                return value
            elif isinstance(value, str):
                try:
                    return datetime.date.fromisoformat(value)
                except ValueError:
                    return value
        
        if field_type == "datetime":
            if isinstance(value, datetime.datetime):
                return value
            elif isinstance(value, datetime.date):
                return datetime.datetime.combine(value, datetime.time.min)
            elif isinstance(value, str):
                try:
                    return datetime.datetime.fromisoformat(value.replace('Z', '+00:00'))
                except ValueError:
                    return value
        
        if field_type in ("decimal", "float"):
            if isinstance(value, Decimal):
                return float(value)
            elif isinstance(value, (int, float)):
                return float(value)
        
        if field_type == "int":
            if isinstance(value, (int, float)):
                return int(value)
        
        if field_type == "bool":
            return "Yes" if value else "No"
        
        return value
    
    def auto_filter(self) -> None:
        max_col = len(self.column_formats)
        if max_col > 0:
            self.ws.auto_filter.ref = f"A1:{get_column_letter(max_col)}{self.current_row - 1}"
    
    def freeze_header(self) -> None:
        self.ws.freeze_panes = "A2"
    
    def save_to_response(self, filename: str = "report") -> HttpResponse:
        buffer = io.BytesIO()
        self.wb.save(buffer)
        buffer.seek(0)
        
        response = HttpResponse(
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}.xlsx"'
        return response


class RunReportExcelView(APIView):
    """Export report directly to Excel."""
    
    permission_classes = [IsAuthenticated]
    
    def post(self, request):
        from .engine import ReportEngine
        from .serializers import ReportConfigSerializer
        
        config_data = request.data.get("config")
        if not config_data:
            return Response(
                {"error": "config is required"},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        serializer = ReportConfigSerializer(data=config_data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        
        # Get all data (no pagination for Excel)
        engine = ReportEngine(config=serializer.validated_data, tenant=request.tenant)
        qs = engine._build_queryset()
        columns = engine._resolved_columns()
        
        # Serialize all rows
        data = []
        for obj in qs:
            data.append(engine._serialize_row(obj, columns))
        
        filename = request.data.get("filename", "report")
        sheet_title = request.data.get("sheet_title", "Report Data")
        
        exporter = ExcelExporter(title=sheet_title)
        exporter.add_header(columns)
        
        for idx, row in enumerate(data):
            exporter.add_row(row, columns, is_alternate=(idx % 2 == 1))
        
        exporter.auto_filter()
        exporter.freeze_header()
        
        return exporter.save_to_response(filename)


class SavedReportExcelView(APIView):
    """Export saved report to Excel."""
    
    permission_classes = [IsAuthenticated]
    
    def post(self, request, report_id):
        from .engine import ReportEngine
        from .models import SavedReport
        
        try:
            report = SavedReport.objects.get(id=report_id, tenant=request.tenant)
        except SavedReport.DoesNotExist:
            return Response({"error": "Report not found"}, status=status.HTTP_404_NOT_FOUND)
        
        engine = ReportEngine(config=report.config, tenant=request.tenant)
        qs = engine._build_queryset()
        columns = engine._resolved_columns()
        
        data = []
        for obj in qs:
            data.append(engine._serialize_row(obj, columns))
        
        exporter = ExcelExporter(title=report.name[:31])
        exporter.add_header(columns)
        
        for idx, row in enumerate(data):
            exporter.add_row(row, columns, is_alternate=(idx % 2 == 1))
        
        exporter.auto_filter()
        exporter.freeze_header()
        
        return exporter.save_to_response(report.name)